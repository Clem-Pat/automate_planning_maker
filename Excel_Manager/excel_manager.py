#import panda
import csv
import os, os.path
import pandas as pd
import openpyxl
import string
import xlwings as xw
import xlsxwriter
from Planning_Manager.creneau import Creneau

"""
L_data = [['Lundi8h', '', '', '', '', '', ''], ['', '', '', '', '', '', ''], ['', 'Mardi12h', '', '', '', 'Samedi12', ''], ['', '', '', '', '', '', ''], ['', '', 'Mercredi16h', '', '', '', ''], ['', '', 'Mercredi18h', 'x', '', '', ''], ['', '', '', '', '', 'Samedi20h', ''], ['', '', '', '', '', '', ''], ['Lundi00h', '', 'Mercredi00h', '', '', '', '']]
"""                     

class Excel_Reader():
    def __init__(self, names, filename=None, type='indispo'):
        self.names = names
        if filename == None:
            path = os.path.dirname(os.path.abspath(__file__))
            self.filename = f"{path[:-16]}/Data/indispo_dispo.xlsx"
        else:
            self.filename = filename
        self.type = type
        self.crens, self.jours = ['12h15-13h', '13h-13h30', '17h30-20h30', '20h30-23h', '23h-00h'], ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
        self.data_of_names = {}
        for name in self.names:
            self.data_of_names[name] = self.extract_data_of_names(name)

    def extract_data_of_names(self, name):
        alph = list(string.ascii_lowercase)
        workBook = openpyxl.load_workbook(filename=self.filename)
        if self.type == 'indispo':
            if len(workBook.sheetnames) > 1: #si le fichier est de la forme "chacun sa sheet, avec un planning en 2D"
                sheet = workBook[name]
                res = [[sheet[str(alph[j])+str(i)].value for j in range(1,len(self.jours)+1)] for i in range(3, len(self.crens)+3)]
                return res
            else:  # si le fichier est de la forme "tous sur la même sheet, avec un planning en 3D"
                sheet = workBook[workBook.sheetnames[0]]
                k = 2
                while k < 20:
                    k+=1
                    if str(sheet[str(alph[0])+str(k)].value) == name: break
                res1 = []
                for j in range(len(self.jours)):
                    res1.append([])
                    for i in range(len(self.crens)):
                        if j*len(self.crens) + i + 1 < len(alph): res1[-1].append(sheet[str(alph[j*len(self.crens) + i + 1]) + str(k)].value)
                        else: res1[-1].append(sheet[str(alph[0]) + str(alph[j*len(self.crens) + i + 1 - 26]) + str(k)].value)
                res = [[0 for i in range(len(self.jours))] for j in range(len(self.crens))]
                for j in range(len(res)):
                    for i in range(len(res[0])):
                        res[j][i] = res1[i][j]
                return res

        elif self.type == 'historic':
            sheet = workBook['Feuil1']
            for i in range(20):
                if str(sheet[str(alph[0])+str(i+2)].value) == name:
                    res = [sheet[str(alph[1])+str(i+2)].value, sheet[str(alph[2])+str(i+2)].value, sheet[str(alph[3])+str(i+2)].value]
                    return res
            return False

        else :
            print("Je ne sais pas lire d'autres fichiers que des historiques ou des indispos", self.filename)

class Excel_Creator():
    def __init__(self, jours, crens, filename=None):
        self.jours, self.crens = jours, crens
        if filename == None:
            self.filename = str(os.path.dirname(os.path.abspath(__file__)))[:-14] + '/Data/Output'
        else:
            self.filename = filename
            if 'Output' not in self.filename[:-7] : self.filename += '/Output'

    def data_from_list_to_dict(self,L_data=None):
        """translates the list of data to dictionnary (useful for excel creation)"""
        res = {"":self.crens}
        for j in range(len(self.jours)):
            res[self.jours[j]] = []
            for i in range(len(self.crens)):
                if isinstance(L_data[i][j], list):
                    try: res[self.jours[j]].append(f'{L_data[i][j][0]}')
                    except : res[self.jours[j]].append('None')
                    for k in range(1, len(L_data[i][j])):
                        if str(L_data[i][j][k]) != 'None' :
                            res[self.jours[j]][-1] += f' & {L_data[i][j][k]}'
                else:
                    res[self.jours[j]].append(L_data[i][j])
        return res

    def create_planning(self, name, L_data, workers=None, caller=None):
        """crée le fichier excel"""
        self.name = name
        self.L_data = L_data
        root = self.filename
        if not os.path.exists(root): os.makedirs(root)
        self.path = root
        try: xw.Book(f'{self.path}/{self.name}.xlsx').close()
        except: pass
        if self.name == 'Publishable_planning':
            self.workers = caller.data.sort_as_default(workers)
            alph = list(string.ascii_uppercase)
            workbook = xlsxwriter.Workbook(f'{self.path}/{self.name}.xlsx')
            worksheet = workbook.add_worksheet()
            for j in range(len(self.jours)):
                if self.jours[j].lower() in caller.data.soirees: message = 'Soirée'
                elif self.jours[j].lower() in caller.data.soirees_mefo: message = 'Mefo'
                else : message = ''
                cell_color = workbook.add_format()
                cell_color.set_bg_color('#a4c2f4')
                worksheet.write(str(alph[3*j + 2]) + str(1), self.jours[j])
                for k in range(1, 4): 
                    if k == 2: worksheet.write(str(alph[3*j + k]) + str(2), message, cell_color)
                    else: worksheet.write(str(alph[3*j + k]) + str(2), "", cell_color)
                for w in range(len(self.workers)):
                    worksheet.write(str(alph[0]) + str(w+3), self.workers[w].name)

                    if self.workers[w].name in caller.data.mefos and self.jours[j].lower() in caller.data.soirees_mefo:
                        mefo_crens_in_person_planning = [cren for cren in self.workers[w].crens if cren.type == 'mefo']
                        if len(mefo_crens_in_person_planning) == 0: # Version 1.3, mefo crens are not counted in the mefo-person's planning. But we still need to render it in the excel. 
                           self.workers[w].crens.insert(0, Creneau(f'{self.jours[j].lower()} mefo'))  
                    
                    crens_of_worker_that_day = [cren for cren in self.workers[w].crens if cren.j == j]
                    #for cren in self.workers[w].crens:
                    #    if cren.j == j: crens_of_worker_that_day.append(cren)
                    if len(crens_of_worker_that_day) == 1:
                        cell_color = workbook.add_format()
                        cell_color.set_bg_color(crens_of_worker_that_day[0].color)
                        for k in range(1, 4):
                            if k == 1: text = crens_of_worker_that_day[0].text
                            else: text = ''
                            worksheet.write(str(alph[3*j+k]) + str(w+3), text, cell_color)
                    elif len(crens_of_worker_that_day) > 1: #If the person has more than one cren in the day
                        for k in range(min(len(crens_of_worker_that_day), 3)):
                            cell_color = workbook.add_format()
                            cell_color.set_bg_color(crens_of_worker_that_day[k].color)
                            worksheet.write(str(alph[3*j+(k+1)]) + str(w+3), crens_of_worker_that_day[k].text, cell_color)
                     
            workbook.close()
        elif self.name == 'Planning_for_each_worker':
            self.workers = workers
            with pd.ExcelWriter(f'{self.path}/{self.name}.xlsx') as writer:
                for i in range(len(self.workers)):
                    caller.refresh_progress_bar(i / (len(caller.planning.workers) + len(caller.names)))
                    data_dic = self.data_from_list_to_dict(self.L_data[i])
                    data_frame = pd.DataFrame(data_dic)
                    data_frame.to_excel(writer, sheet_name=self.workers[i].name, index=False)
        elif self.name == 'General_planning':
            data_dic = self.data_from_list_to_dict(L_data=self.L_data)
            data_frame = pd.DataFrame(data_dic)
            data_frame.to_excel(f'{self.path}/{self.name}.xlsx', sheet_name='Sheet1', index=False)
        try : xw.Book(f'{self.path}/{self.name}.xlsx').close()
        except: pass

class Excel_Modifier():
    def __init__(self, path):
        self.path = path

    def get_value(self, case, sheet_name=None):
        wb = openpyxl.load_workbook(self.path)
        if sheet_name != None: ws = wb[sheet_name]
        else : ws = wb.active
        resu = str(ws[case].value)
        wb.save(self.path)
        return resu

    def modify_case(self, case, value, sheet_name=None):
        wb = openpyxl.load_workbook(self.path)
        if sheet_name != None: ws = wb[sheet_name]
        else: ws = wb.active
        ws[case] = value
        wb.save(self.path)
        return True

if __name__ == '__main__':
    names = ['Alice']
    indispo_dispo_excel = Excel_Reader(names, '../Data/indispo_dispo.xlsx')
    excel = Excel_Creator(['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche'], ['12h15-13h', '13h-13h30', '17h30-20h30', '20h30-23h', '23h-00h'])
    print(indispo_dispo_excel.data_of_names['Alice'])
    excel.create_planning('Alice', indispo_dispo_excel.data_of_names['Alice'])
